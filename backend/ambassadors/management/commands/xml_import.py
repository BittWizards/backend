import re
from datetime import datetime

import django.apps
from django.core.management.base import BaseCommand, CommandError
from django.db import models, transaction
from django.utils.timezone import make_aware
from openpyxl.reader.excel import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from ambassadors_project.constants import REGEX_PATTERN


class Command(BaseCommand):
    def add_arguments(self, parser):
        """Регистрируем ключи запуска."""
        parser.add_argument("-f", "--file", type=str, help="Файл для импорта")
        parser.add_argument(
            "-m",
            "--manual",
            default=False,
            action="store_true",
            help=" Выключить автоподбор моделей для таблиц.",
        )

    def handle(self, *args, **kwargs):
        """Начало команды."""
        self.models = django.apps.apps.get_models()
        self.imported = []

        try:
            if file := kwargs.get("file"):
                self.sheets = load_workbook(file)
            else:
                self.sheets = load_workbook("data/test_data.xlsx")
        except Exception as error:
            raise CommandError(f"Ошибка при загрузке файла: {error}")

        self.get_models_relations()
        if kwargs.get("manual"):
            self.run_comparison()
        else:
            self.run_auto_comparison()

        try:
            with transaction.atomic():
                self.choose_import_model()
        except Exception as error:
            raise CommandError(f"Сбой при импорте: {error}")

        self.stdout.write(self.style.SUCCESS("Импорт прошел успешно"))

    def get_models_relations(self):
        """Формируем словарь зависимостей для моделей."""
        self.relations = {}
        for model in self.models:
            self.relations[model] = []
            for field in model._meta._get_fields(reverse=False):
                if field.is_relation:
                    related_model = field.related_model
                    self.relations[model].append(related_model)

    def run_auto_comparison(self):
        """Формируем словарь, где для каждой таблицы сопоставляем модель."""
        self.comparison = {}
        for sheet in self.sheets:
            model = self.find_model_by_sheet_name(sheet.title)
            if not model:
                fields_names = tuple(sheet.iter_rows(1, 1, values_only=True))[
                    0
                ]
                model = self.find_model_by_fields_names(
                    fields_names, sheet.title
                )
            self.comparison[sheet] = model

    def run_comparison(self):
        """
        Просим пользователя сформировать словарь,
        где каждой таблице соответствует модель.
        """
        self.comparison = {}
        for sheet in self.sheets:
            while True:
                self.stdout.write(
                    f"Выберите модель для таблицы {sheet.title}:"
                )
                for index, model in enumerate(self.models):
                    self.stdout.write(f"{index + 1}. {model.__qualname__}")
                try:
                    number = int(
                        input(
                            "Введите число соответствующее номеру "
                            "модели из предложенного списка или 0, "
                            "если эту таблицу импортировать не нужно: "
                        )
                    )
                    self.models[number - 1]
                    if number > 0:
                        self.comparison[sheet] = self.models[number - 1]
                    break
                except Exception:
                    self.stdout.write(
                        self.style.WARNING("Введены неверные данные.")
                    )

    def choose_import_model(self):
        """Импортируем данные, учитывая порядок для импорта."""
        comparison = self.comparison.copy()
        while True:
            if len(self.comparison) == len(self.imported):
                break
            comparison_copy = comparison.copy()
            for sheet, model in comparison_copy.items():
                if all(
                    item in self.imported for item in self.relations[model]
                ):
                    self.import_models(sheet, model)
                    self.imported.append(model)
                    comparison.pop(sheet)

    def import_models(self, obj: Worksheet, model: models.Model):
        """Импорт данных с листа excel."""
        fields_names = tuple(obj.iter_rows(1, 1, values_only=True))[0]
        cells = obj.iter_rows(2, values_only=True)

        for values in cells:
            values_dict = {}
            many_to_many_values_dict = {}
            for index, field_name in enumerate(fields_names):
                if model._meta.get_field(field_name).many_to_many:
                    many_to_many_values_dict[field_name] = values[index]
                else:
                    value = values[index]
                    if isinstance(value, datetime):
                        value = make_aware(value)
                    values_dict[field_name] = value

            model_obj, _ = model.objects.get_or_create(**values_dict)

            for field_name in many_to_many_values_dict.keys():
                related_manager = getattr(model_obj, field_name)
                related_manager.set(
                    str(many_to_many_values_dict[field_name]).split(", ")
                )

        self.stdout.write(
            self.style.SUCCESS(f"Данные для {model.__qualname__} загружены")
        )

    def find_model_by_sheet_name(self, sheet_name: str):
        """Поиск модели по названию листа."""
        for model in self.models:
            pattern_for_sheet = re.compile(
                REGEX_PATTERN.format(name=model.__qualname__.lower())
            )
            if pattern_for_sheet.match(sheet_name.lower()):
                return model
            pattern_for_model = re.compile(
                REGEX_PATTERN.format(name=sheet_name.lower())
            )
            if pattern_for_model.match(model.__qualname__.lower()):
                return model

    def find_model_by_fields_names(
        self, fields_names: list[str], sheet_name: str
    ):
        """Поиск подходящей модели по названиям полей."""
        suitable_models = []
        for model in self.models:
            model_fields_names = {
                field.name for field in model._meta._get_fields(reverse=False)
            }

            if all(field in model_fields_names for field in fields_names):
                suitable_models.append(model)
            else:
                if self._check_fields_with_postfix(
                    model_fields_names, fields_names
                ):
                    suitable_models.append(model)
        if len(suitable_models) > 1:
            return self._ask_for_choose(suitable_models, sheet_name)
        elif len(suitable_models) == 1:
            return suitable_models[0]
        fields_for_error = ", ".join([x for x in fields_names])
        raise CommandError(
            f"Не удалось найти модель для листа {sheet_name} "
            f"с полями {fields_for_error}."
        )

    def _check_fields_with_postfix(
        self,
        model_fields_names: tuple[str],
        fields_names: list[str],
        postfix="_id",
    ):
        """Проверка наличия всех полей в модели с учетом постфикса."""
        id_field_pattern = re.compile(rf"\b\w+{postfix}\b")
        id_fields = [
            field for field in fields_names if id_field_pattern.match(field)
        ]
        non_id_fields = set(fields_names) - set(id_fields)
        if all(
            field.replace("_id", "") in model_fields_names
            for field in id_fields
        ):
            return all(field in model_fields_names for field in non_id_fields)

    def _ask_for_choose(self, models_choice: list[models.Model], sheet_name):
        """Просим пользователя выбрать подходящую модель."""
        for index, model in enumerate(models_choice):
            self.stdout.write(f"{index + 1}. {model.__qualname__}")
        while True:
            try:
                num = int(
                    input(
                        f"Выберите модель соотвествующую листу {sheet_name}: "
                    )
                )
                if num < 1:
                    raise Exception
                models_choice[num - 1]
                break
            except Exception:
                self.stdout.write(
                    self.style.WARNING("Введены неверные данные.")
                )
        return models_choice[num - 1]
